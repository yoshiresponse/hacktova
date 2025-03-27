import pandas as pd
from openpyxl import Workbook

def group_processed_sheet(input_file, output_file):
    width = 1  # Number of columns in the group path (H, I, J)

    # Read the 'Processed' worksheet without headers and skip the first row.
    df = pd.read_excel(input_file, sheet_name='Processed', header=None, skiprows=1)

    # Pre-clean the grouping columns (H, I, J => indices 7, 8, 9)
    # Convert non-null values to stripped strings,
    # and fill nulls with an empty string so that only H is used to check for emptiness.
    for col in [7, 8, 9]:
        df[col] = df[col].apply(lambda x: str(x).strip() if pd.notna(x) else "")

    # Group rows by the three path columns (columns H, I, J are indices 7, 8, 9)
    groups = df.groupby([7, 8, 9], as_index=False)

    # Create a new workbook (remove the automatic sheet)
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Create a dictionary to track parent groups (path_H with an empty path_I)
    # Keys in this dictionary are the value of column H.
    # Each value holds the parent's worksheet, its group, and a list of subgroup sheet names.
    parent_map = {}

    # Loop through each group to create sheets
    for keys, group in groups:
        # keys is a tuple with three path values (columns H, I, J)
        path_H, path_I, path_J = keys
        print(path_H, path_I, path_J)

        # Only create a group if path_H is nonempty (I and J can be empty)
        if not path_H:
            continue

        # Sort the group rows by column D (index 3) in ascending alphabetical order.
        group = group.sort_values(by=[10])

        # Use the first row's value in column C (index 2) as the sheet name.
        sheet_name = str(group.iloc[0, 2])
        # Ensure the sheet name is a valid Excel name (max 31 characters).
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]

        ws = wb.create_sheet(title=sheet_name)

        # Write the group path in first row cells A, B, C.
        ws['A1'] = path_H
        ws['B1'] = path_I
        ws['C1'] = path_J

        # Starting from column D (column index 4), process each row in the group.
        # For each row, write two consecutive cells:
        #   first cell (col2): value from column K (index 10)
        #   second cell (below in row 2): value from column Q (index 16)
        start_col = 4  # Column D
        for idx, (_, row) in enumerate(group.iterrows()):
            offset = idx * width
            col1 = start_col + offset      # (unused cell if re-enabled for row[3])
            col2 = col1 + 1                # for column K / Q values.
            # ws.cell(row=1, column=col1, value=row[3])
            ws.cell(row=1, column=col2, value=row[10])
            ws.cell(row=2, column=col2, value=row[16])

        # If this group is a parent (I is empty), store its worksheet and group info.
        if path_I == "":
            parent_map[path_H] = {"sheet": ws, "group": group, "subgroups": []}
        else:
            # If I is not empty, check whether there's a parent with the same path_H.
            if path_H in parent_map:
                parent_map[path_H]["subgroups"].append(sheet_name)

    # After processing all groups, update each parent's worksheet with its subgroups.
    for parent_val, info in parent_map.items():
        ws = info["sheet"]
        group = info["group"]
        subgroups = info["subgroups"]
        if subgroups:
            # Determine the next column following the group's last block.
            # Each row in the group produced a three-cell block starting in column D.
            num_blocks = group.shape[0]
            # Last used column = start_col + (num_blocks-1)*3 + 2 (since each block takes 3 columns)
            last_used = 4 + (num_blocks - 1) * width + 2 if num_blocks > 0 else width
            next_col = last_used + 1
            # Write a comma-separated list of subgroup sheet names prefixed with "Subgroups:".
            ws.cell(row=1, column=next_col, value="Subgroups: " + ", ".join(subgroups))

    wb.save(output_file)
    print(f"Created grouped Excel file: {output_file}")

if __name__ == "__main__":
    input_excel = "LibE2025dev.xlsx"      # Input file
    output_excel = "grouped_output_with_values_compact.xlsx"  # Output file
    group_processed_sheet(input_excel, output_excel)